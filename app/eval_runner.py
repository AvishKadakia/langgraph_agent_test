"""Core eval engine: run Excel test cases against a live agent + LLM judge.

Self-contained (no dependency on the agent repo). Exposes ``run_eval`` which
streams progress through a callback so the web UI can show live results, and
writes a styled results workbook.
"""
from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import subprocess
import traceback
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urlparse

import requests
from openai import AzureOpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

INCIDENT_RE = re.compile(r"\b(?:INC|REQ|CHG|RITM|PRB|TASK|CASE)\d+\b", re.IGNORECASE)


# --------------------------------------------------------------------------- #
# Config / clients
# --------------------------------------------------------------------------- #
def env(name: str, default: str | None = None) -> str | None:
    return os.environ.get(name) or default


def get_agent_token() -> str:
    """Bearer token for the agent API: CHAT_API_TOKEN if set, else az login."""
    explicit = os.getenv("CHAT_API_TOKEN")
    if explicit:
        return explicit
    from azure.identity import AzureCliCredential

    scope = env("CHAT_API_SCOPE")
    if not scope:
        raise RuntimeError("CHAT_API_SCOPE is not set (and no CHAT_API_TOKEN provided).")
    return AzureCliCredential().get_token(scope).token


# --------------------------------------------------------------------------- #
# Agent auth
#
# The deployed agent is fronted by the `agent-web-ui` app, which uses Auth.js
# (NextAuth) session cookies — NOT bearer tokens — and exposes the agent on its
# BFF route `POST /api/chat`. So we authenticate "like the browser": send the
# session cookie and call `/api/chat`. The cookie is sourced from CHAT_API_COOKIE
# (set this when running in the container) or auto-read from the local Chrome
# profile. As a fallback, if no cookie is available we use a bearer token against
# `/chat` (for a backend that speaks that protocol directly).
# --------------------------------------------------------------------------- #
def resolve_agent_auth(base_url: str) -> dict[str, Any]:
    cookie = os.getenv("CHAT_API_COOKIE") or _extract_browser_cookie(base_url)
    if cookie:
        return {"mode": "cookie", "headers": {"Cookie": cookie}, "chat_path": "/api/chat"}
    return {
        "mode": "bearer",
        "headers": {"Authorization": f"Bearer {get_agent_token()}"},
        "chat_path": "/chat",
    }


def _extract_browser_cookie(base_url: str) -> str | None:
    """Best-effort: read the session cookie(s) for ``base_url``'s host from the
    local Chrome cookie store (macOS) and return them as a ``Cookie`` header.

    Returns None if Chrome / the Keychain key aren't available (e.g. inside the
    container) — set ``CHAT_API_COOKIE`` in that case. Any failure degrades to
    None rather than raising, so the bearer fallback still applies."""
    host = urlparse(base_url).hostname
    if not host:
        return None
    db = Path.home() / "Library/Application Support/Google/Chrome/Default/Cookies"
    if not db.exists():
        return None
    # Retry: copying the (locked) cookie DB or the Keychain lookup can transiently
    # fail. A silent bearer fallback against a cookie-only host would 405, so it is
    # worth a couple of attempts before giving up.
    last_exc: Exception | None = None
    for _attempt in range(3):
        try:
            return _read_chrome_cookie(db, host)
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
    _ = last_exc  # extraction unavailable (e.g. in the container) — caller falls back
    return None


def _read_chrome_cookie(db: Path, host: str) -> str | None:
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.primitives.hashes import SHA1
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

    pw = subprocess.run(
        ["security", "find-generic-password", "-ws", "Chrome Safe Storage"],
        capture_output=True, text=True, timeout=20,
    ).stdout.strip().encode()
    if not pw:
        raise RuntimeError("empty Chrome Safe Storage key")
    key = PBKDF2HMAC(
        algorithm=SHA1(), length=16, salt=b"saltysalt",
        iterations=1003, backend=default_backend(),
    ).derive(pw)

    tmp = Path("/tmp/.eval_chrome_cookies.db")
    shutil.copy(db, tmp)
    con = sqlite3.connect(tmp)
    try:
        rows = con.execute(
            "SELECT name, encrypted_value FROM cookies WHERE host_key = ?", (host,)
        ).fetchall()
    finally:
        con.close()

    def _decrypt(ev: bytes) -> str | None:
        if not ev:
            return None
        if ev[:3] in (b"v10", b"v11"):
            ev = ev[3:]
        dec = Cipher(
            algorithms.AES(key), modes.CBC(b" " * 16), backend=default_backend()
        ).decryptor()
        pt = dec.update(ev) + dec.finalize()
        pt = pt[: -pt[-1]]  # strip PKCS#7 padding
        try:
            return pt.decode("utf-8")
        except UnicodeDecodeError:
            return pt[32:].decode("utf-8", "ignore")  # newer Chrome prepends a 32-byte domain hash

    pairs = []
    for name, ev in rows:
        val = _decrypt(ev)
        if val:
            pairs.append(f"{name}={val}")
    return "; ".join(pairs) or None


def judge_client() -> AzureOpenAI:
    """Azure OpenAI client for the judge. Uses an API key if provided, otherwise
    falls back to the same az login (AAD token) — no secret required."""
    endpoint = env("AZURE_OPENAI_ENDPOINT")
    if not endpoint:
        raise RuntimeError("AZURE_OPENAI_ENDPOINT is not set.")
    version = env("AZURE_OPENAI_API_VERSION", "2024-08-01-preview")
    key = env("AZURE_OPENAI_API_KEY")
    if key:
        return AzureOpenAI(azure_endpoint=endpoint, api_key=key, api_version=version)
    from azure.identity import AzureCliCredential, get_bearer_token_provider

    provider = get_bearer_token_provider(
        AzureCliCredential(), "https://cognitiveservices.azure.com/.default"
    )
    return AzureOpenAI(azure_endpoint=endpoint, azure_ad_token_provider=provider, api_version=version)


# --------------------------------------------------------------------------- #
# Excel case loading
# --------------------------------------------------------------------------- #
def _cell_to_json_value(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        if value.lower() == "yes":
            return True
        if value.lower() == "no":
            return False
        if value.upper() in {"N/A", "NA"}:
            return None
    return value


def _parse_incident_numbers(value: Any) -> list[str]:
    if not value:
        return []
    return sorted({m.upper() for m in INCIDENT_RE.findall(str(value))})


def load_cases(excel_path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() for h in rows[0]]
    cases: list[dict[str, Any]] = []
    for raw in rows[1:]:
        row = {h: _cell_to_json_value(v) for h, v in zip(headers, raw)}
        if not row.get("Test ID") or not row.get("User Query"):
            continue
        row["Expected Incident Numbers Parsed"] = _parse_incident_numbers(
            row.get("Expected Incident Numbers Returned")
        )
        cases.append(row)
    return cases


def list_sheets(excel_path: Path) -> list[str]:
    return load_workbook(excel_path, read_only=True).sheetnames


# --------------------------------------------------------------------------- #
# Agent call
# --------------------------------------------------------------------------- #
def call_chat_route(base_url: str, auth: dict[str, Any], case: dict[str, Any]) -> dict[str, Any]:
    test_id = case["Test ID"]
    resp = requests.post(
        f"{base_url}{auth['chat_path']}",
        headers=auth["headers"],
        json={
            "message": case["User Query"],
            "session_id": f"eval-{test_id}",
            "system_prompt": (
                "Answer the user query using the available tools. Be concise. "
                "Include relevant ServiceNow incident numbers when applicable. "
                "Do not invent incident numbers."
            ),
        },
        timeout=240,
        allow_redirects=False,  # a redirect means auth failed (e.g. bounced to /api/auth/login)
    )
    if resp.status_code != 200:
        raise RuntimeError(_chat_error(resp, auth))
    return resp.json()


def _chat_error(resp: requests.Response, auth: dict[str, Any]) -> str:
    """Build an error message, adding a hint when a cookie session looks expired."""
    msg = f"HTTP {resp.status_code}: {resp.text[:500]}"
    redirected_to_login = "auth/login" in resp.headers.get("location", "")
    if auth["mode"] == "cookie" and (resp.status_code in (401, 403) or 300 <= resp.status_code < 400 or redirected_to_login):
        msg += (
            " — the browser session cookie is missing/expired. Log in to the web UI "
            "again, then set CHAT_API_COOKIE (or re-run locally to auto-read Chrome)."
        )
    return msg


# --------------------------------------------------------------------------- #
# Judge
# --------------------------------------------------------------------------- #
JUDGE_SCHEMA = {
    "name": "servicenow_excel_eval_judgement",
    "strict": True,
    "schema": {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "verdict": {"type": "string", "enum": ["PASS", "FAIL"]},
            "score": {"type": "integer", "minimum": 0, "maximum": 100},
            "reason": {"type": "string"},
            "missing_requirements": {"type": "array", "items": {"type": "string"}},
            "incorrect_or_unsupported_claims": {"type": "array", "items": {"type": "string"}},
        },
        "required": [
            "verdict", "score", "reason", "missing_requirements",
            "incorrect_or_unsupported_claims",
        ],
    },
}

JUDGE_SYSTEM_PROMPT = (
    "You are a strict QA judge for a ServiceNow + knowledge-base agent.\n\n"
    "You will receive one Excel test case converted to JSON and the actual agent "
    "response from the /chat route.\n\n"
    "Judge PASS only when the response satisfies the row's expected behavior. Use the "
    "whole row, not just Expected Output Summary.\n\n"
    "Primary criteria:\n"
    "1. The response answers the User Query.\n"
    "2. The response matches the Expected Intent.\n"
    "3. The response is consistent with Route To.\n"
    "4. The response respects Data Source (Resolved).\n"
    "5. If STTM Lookup Required? is true, the answer should reflect that lookup.\n"
    "6. If Expected Incident Numbers Parsed is non-empty, all expected incident numbers must be present.\n"
    "7. The response must not invent unsupported incident numbers.\n"
    "8. Do not fail for minor wording or formatting differences.\n"
    "9. Fail if the answer is generic, evasive, contradicts the expected row, or omits required incidents.\n\n"
    "Return only the structured JSON verdict."
)


def judge_agent_output(client: AzureOpenAI, model: str, case: dict[str, Any], chat_response: dict[str, Any]) -> dict[str, Any]:
    judge_input = {
        "test_case": case,
        "agent_response": {
            "answer": chat_response.get("answer"),
            "errors": chat_response.get("errors", []),
            "metadata": chat_response.get("metadata", {}),
        },
    }
    completion = client.chat.completions.create(
        model=model,
        response_format={"type": "json_schema", "json_schema": JUDGE_SCHEMA},
        messages=[
            {"role": "system", "content": JUDGE_SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(judge_input, indent=2, default=str)},
        ],
    )
    content = completion.choices[0].message.content
    assert content, "Judge returned empty content"
    return json.loads(content)


def failure_judgement(stage: str, error: BaseException) -> dict[str, Any]:
    return {
        "verdict": "FAIL", "score": 0,
        "reason": f"{stage} failed: {type(error).__name__}: {error}",
        "missing_requirements": [], "incorrect_or_unsupported_claims": [str(error)],
    }


# --------------------------------------------------------------------------- #
# Self-diagnosis (debug aid on FAIL)
# --------------------------------------------------------------------------- #
DIAGNOSIS_SYSTEM_PROMPT = (
    "You are in DEBUG mode. Explain, in plain and simple language, why a previous answer failed "
    "a QA check. Be brief and concrete. Do NOT re-answer the question, do NOT output ticket tables, "
    "and do NOT add anything beyond the requested 3 lines."
)


def ask_agent_why_failed(base_url: str, auth: dict[str, Any], case: dict[str, Any], chat_response: dict[str, Any], judgement: dict[str, Any]) -> str:
    prompt = (
        "DEBUG: your previous answer failed a QA check. Explain why, briefly and in plain language.\n\n"
        f"Question asked: {case.get('User Query')}\n"
        f"Expected incident(s): {case.get('Expected Incident Numbers Parsed')}\n"
        f"Expected outcome: {case.get('Expected Output Summary')}\n"
        f"Why the grader rejected your answer: {judgement.get('reason')}\n\n"
        "Reply in EXACTLY these 3 lines, one short sentence each (max ~25 words per line), "
        "no tables, no preamble:\n"
        "Root cause: <the single main reason it failed>\n"
        "What happened: <what your answer did vs. what was expected, e.g. which incidents were missed or wrongly added>\n"
        "Fix: <the one change needed to pass>"
    )
    resp = requests.post(
        f"{base_url}{auth['chat_path']}",
        headers=auth["headers"],
        json={"message": prompt, "session_id": f"eval-{case['Test ID']}", "system_prompt": DIAGNOSIS_SYSTEM_PROMPT},
        timeout=240,
        allow_redirects=False,
    )
    if resp.status_code != 200:
        return f"[diagnosis call failed: HTTP {resp.status_code}: {resp.text[:300]}]"
    return resp.json().get("answer", "") or ""


# --------------------------------------------------------------------------- #
# Per-case runner
# --------------------------------------------------------------------------- #
def run_case(base_url: str, auth: dict[str, Any], client: AzureOpenAI, judge_model: str, case: dict[str, Any]) -> dict[str, Any]:
    chat_response: dict[str, Any] | None = None
    failure_stage: str | None = None
    exception: BaseException | None = None
    try:
        failure_stage = "chat_route"
        chat_response = call_chat_route(base_url, auth, case)
        failure_stage = "llm_judge"
        judgement = judge_agent_output(client, judge_model, case, chat_response)
        failure_stage = None
    except Exception as exc:  # noqa: BLE001
        exception = exc
        judgement = failure_judgement(failure_stage or "unknown", exc)

    diagnosis = ""
    if judgement.get("verdict") == "FAIL" and chat_response is not None and exception is None:
        try:
            diagnosis = ask_agent_why_failed(base_url, auth, case, chat_response, judgement)
        except Exception as exc:  # noqa: BLE001
            diagnosis = f"[diagnosis error: {type(exc).__name__}: {exc}]"

    return build_result_row(case, chat_response, judgement, failure_stage, exception, diagnosis)


def build_result_row(case, chat_response, judgement, failure_stage, exception, diagnosis="") -> dict[str, Any]:
    chat_response = chat_response or {}
    return {
        "Run Timestamp UTC": datetime.now(timezone.utc).isoformat(),
        "Test ID": case.get("Test ID"),
        "Judge Verdict": judgement.get("verdict"),
        "Judge Score": judgement.get("score"),
        "Failure Stage": failure_stage or "",
        "User Query": case.get("User Query"),
        "Expected Intent": case.get("Expected Intent"),
        "Route To": case.get("Route To"),
        "Data Source (Resolved)": case.get("Data Source (Resolved)"),
        "STTM Lookup Required?": case.get("STTM Lookup Required?"),
        "Expected Incident Numbers Parsed": case.get("Expected Incident Numbers Parsed"),
        "Expected Incident Numbers Returned": case.get("Expected Incident Numbers Returned"),
        "Expected Output Summary": case.get("Expected Output Summary"),
        "Agent Answer": chat_response.get("answer"),
        "Judge Reason": judgement.get("reason"),
        "Agent Failure Reasoning": diagnosis,
        "Missing Requirements": judgement.get("missing_requirements", []),
        "Incorrect or Unsupported Claims": judgement.get("incorrect_or_unsupported_claims", []),
        "Chat Errors": chat_response.get("errors", []),
        "Request ID": chat_response.get("request_id"),
        "Session ID": chat_response.get("session_id"),
        "Exception": str(exception) if exception else "",
        "Traceback": "".join(traceback.format_exception(exception)) if exception else "",
    }


# --------------------------------------------------------------------------- #
# Results workbook
# --------------------------------------------------------------------------- #
RESULT_HEADERS = [
    "Run Timestamp UTC", "Test ID", "Judge Verdict", "Judge Score", "Failure Stage",
    "User Query", "Expected Intent", "Route To", "Data Source (Resolved)",
    "STTM Lookup Required?", "Expected Incident Numbers Parsed",
    "Expected Incident Numbers Returned", "Expected Output Summary", "Agent Answer",
    "Judge Reason", "Agent Failure Reasoning", "Missing Requirements",
    "Incorrect or Unsupported Claims", "Chat Errors", "Request ID", "Session ID",
    "Exception", "Traceback",
]
WIDE_COLUMNS = {
    "Agent Answer", "Judge Reason", "Agent Failure Reasoning", "Expected Output Summary",
    "Missing Requirements", "Incorrect or Unsupported Claims", "Traceback",
}


def _json_for_excel(value: Any) -> str:
    if value is None:
        return ""
    text = value if isinstance(value, str) else json.dumps(value, indent=2, default=str)
    return text[:32000] + "\n...[truncated]" if len(text) > 32000 else text


def _style_sheet(ws: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = header_fill, header_font, wrap
    for row in ws.iter_rows(min_row=2):
        vcell = None
        for cell in row:
            cell.alignment = wrap
            if ws.cell(row=1, column=cell.column).value == "Judge Verdict":
                vcell = cell
        if vcell and vcell.value == "PASS":
            vcell.fill = pass_fill
        elif vcell and vcell.value == "FAIL":
            vcell.fill = fail_fill
    for col in ws.columns:
        header = str(col[0].value or "")
        max_len = max((len(str(c.value or "")) for c in col[:100]), default=12)
        ws.column_dimensions[col[0].column_letter].width = 60 if header in WIDE_COLUMNS else min(max(max_len + 2, 12), 35)


def _append_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(RESULT_HEADERS)
    for row in rows:
        ws.append([_json_for_excel(row.get(h)) for h in RESULT_HEADERS])
    _style_sheet(ws)


def write_results_excel(results, out_path: Path, excel_name: str, sheet: str, base_url: str) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    total = len(results)
    passed = sum(1 for r in results if r.get("Judge Verdict") == "PASS")
    for r in [
        ("Run Timestamp UTC", datetime.now(timezone.utc).isoformat()),
        ("Target URL", base_url),
        ("Source Excel", excel_name),
        ("Source Sheet", sheet),
        ("Total Cases", total),
        ("Passed", passed),
        ("Failed", total - passed),
        ("Pass Rate", f"{(passed / total * 100):.2f}%" if total else "0.00%"),
    ]:
        summary.append(r)
    for i in range(1, 9):
        summary[f"A{i}"].font = Font(bold=True)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 90
    _append_sheet(wb, "All Results", results)
    _append_sheet(wb, "PASS", [r for r in results if r.get("Judge Verdict") == "PASS"])
    _append_sheet(wb, "FAIL", [r for r in results if r.get("Judge Verdict") == "FAIL"])
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Orchestrator
# --------------------------------------------------------------------------- #
def run_eval(
    excel_path: Path,
    results_dir: Path,
    *,
    sheet: str | None = None,
    base_url: str | None = None,
    limit: int | None = None,
    progress_cb: Callable[[dict[str, Any]], None] | None = None,
) -> dict[str, Any]:
    """Run all cases sequentially. ``progress_cb`` is called with an event dict
    after setup and after each case. Returns the final summary."""
    base_url = (base_url or env("CHAT_API_URL") or "").rstrip("/")
    if not base_url:
        raise RuntimeError("CHAT_API_URL is not set.")
    sheet = sheet or env("DEFAULT_SHEET", "Intent Test Cases")

    cases = load_cases(excel_path, sheet)
    if limit:
        cases = cases[:limit]

    auth = resolve_agent_auth(base_url)
    client = judge_client()
    judge_model = env("AZURE_OPENAI_CHAT_DEPLOYMENT")
    if not judge_model:
        raise RuntimeError("AZURE_OPENAI_CHAT_DEPLOYMENT is not set.")

    if progress_cb:
        progress_cb({"type": "start", "total": len(cases), "url": base_url, "sheet": sheet})

    results: list[dict[str, Any]] = []
    # Sequential on purpose: the live agent currently 500s under parallel load.
    for i, case in enumerate(cases, 1):
        row = run_case(base_url, auth, client, judge_model, case)
        results.append(row)
        if progress_cb:
            progress_cb({"type": "case", "index": i, "total": len(cases), "row": row})

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = results_dir / f"results_{excel_path.stem}_{stamp}.xlsx"
    write_results_excel(results, out_path, excel_path.name, sheet, base_url)

    passed = sum(1 for r in results if r.get("Judge Verdict") == "PASS")
    summary = {
        "type": "done",
        "total": len(results),
        "passed": passed,
        "failed": len(results) - passed,
        "pass_rate": round(passed / len(results) * 100, 1) if results else 0.0,
        "result_file": out_path.name,
    }
    if progress_cb:
        progress_cb(summary)
    return summary
